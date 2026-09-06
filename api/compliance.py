"""
HUJJATLAR TO'LIQLIGI CHEKLISTI (REJA.md P0-8)
=============================================
"Bu tenderga ariza berish uchun qaysi hujjatlar kerak va ular bizda bormi?"

MVP CHEKLOVI — ONGLI SODDALASHTIRISH:
    Bu STATIK CHEKLIST. Qoidaga asoslangan: tender matnidan hujjat TALABLARI
    qidiriladi va kompaniya bazasidagi hujjatlar bilan solishtiriladi.
    Hujjat MAZMUNINING huquqiy to'g'riligi TEKSHIRILMAYDI va bu ATAYIN
    shunday — aks holda mahsulot noto'g'ri huquqiy kafolat hissini yaratardi.
    Modul HECH QANDAY AI/model chaqirmaydi.

UCHTA MANBA:
    1) BAZAVIY RO'YXAT — biznes-jarayonda (oddiy xarid, 10-11 bosqich)
       buyurtmachi namunaviy ravishda so'raydigan 6 ta hujjat. Ular tender
       matnida yozilmagan bo'lsa ham cheklistda turadi, chunki ariza
       to'plamining odatiy tarkibi shu.
    2) TENDER MATNI — anno, method_marks, tovar spetsifikatsiyasi va
       (mavjud bo'lsa) hujjatlardan ajratilgan matn. Topilgan har talab
       DALIL (evidence) bilan ko'rsatiladi: matnning aynan qaysi bo'lagidan
       kelib chiqqani. Bu NFR shaffoflik talabi.
    3) KOMPANIYA BAZASI — `company_document` (schema_patch_compliance.sql).

HALOL HOLAT: tender matnida hech narsa topilmasa cheklist buni OCHIQ aytadi
    (`detected_from_tender = false`, xabar bilan). O'lchov: 342 tenderning
    aksariyatida anno qisqa (o'rtacha ~132 belgi) va hujjat talablari
    umuman yozilmagan — ular biriktirilgan PDF ichida qoladi. Bo'sh ro'yxat
    ko'rsatib "hujjat kerak emas" degan taassurot qoldirish XATO bo'lardi.

ALIFBO: manba matni aralash (ruscha kirill, o'zbek lotin, o'zbek kirill).
    Barcha qidiruv api/translit.py orqali — o'zak bitta alifboda yozilsa ham
    ikkalasi topiladi ("сертификат" <-> "sertifikat"). Alifbo bilan
    tarjima qilib bo'lmaydigan so'zlar (свидетельство <-> guvohnoma)
    naqshlarда ALOHIDA yoziladi.

ANIQLIK (precision) ustuvor: yalang'och kalit so'z juda ko'p soxta natija
    beradi. O'lchov (342 ta tender, 2113 pozitsiya):
        "соответстви" -> 58 ta tender, deyarli hammasi "в соответствии с ..."
        "налогов"     -> 20 ta, ko'pi "анаЛОГОВых" ichidan (fold qurboni)
        "аккредит"    -> "аккредитива" (akkreditiv, akkreditatsiya emas)
    Shuning uchun qoida bitta so'z emas, O'ZAKLAR YAQINLIGI + ISTISNOLAR.
"""
import csv
import datetime as _dt
import io
import re
from functools import lru_cache
from typing import Any, Dict, List, Optional, Sequence, Tuple

from api import translit

#: Ko'p o'zakli qoidada o'zaklar orasidagi eng katta masofa (belgi).
#: 80 ~ bitta jumla bo'lagi. Kattaroq qilinса "сертификат" bir jumlada,
#: "соответстви" boshqasida bo'lsa ham mos kelib qolardi.
WINDOW = 80

#: Dalil (evidence) uchun ajratib olinadigan matn bo'lagi uzunligi.
EVIDENCE_PAD = 70

#: Muddati tugashiga shuncha kun qolganda "tugayapti" deb ogohlantiramiz.
EXPIRING_SOON_DAYS = 30

#: BIZNES VAQT MINTAQASI — O'zbekiston, UTC+5, yozgi vaqt YO'Q.
#:
#: NEGA ANIQ YOZILGAN (o'lchangan 2026-08-30): `doc_status()`,
#: `_days_left()` va `build_checklist()` `datetime.date.today()` ni
#: ishlatardi — u JARAYONNING vaqt mintaqasiga qarab ishlaydi. Baza
#: esa `Asia/Tashkent` da yuradi (`current_setting('TimeZone')` bilan
#: tekshirilgan).
#:
#: Server UTC da yursa (bulutda odatiy holat) har kuni 19:00–24:00
#: UTC oralig'ida jarayonning sanasi bazanikidan BIR KUN ORQADA
#: bo'ladi. Natijada:
#:   * bugun tugaydigan hujjat "ertaga tugaydi" bo'lib ko'rinardi;
#:   * kecha tugagan hujjat "bugun tugaydi" — ya'ni `expired`
#:     o'rniga `expiring_soon`;
#:   * sinovlar kunning qaysi soatida yurgizilganiga qarab goh
#:     o'tib, goh yiqilardi (VAQT MINTAQASIGA BOG'LIQ MO'RTLIK).
#:
#: `etl_tenders.py` da xuddi shu qaror allaqachon qabul qilingan
#: (`TZ = timezone(timedelta(hours=5))`) — bu yerda u takrorlanadi,
#: chunki ikkala modul ham mustaqil ishlatiladi.
BIZNES_TZ = _dt.timezone(_dt.timedelta(hours=5))


def bugun() -> _dt.date:
    """BIZNES kuni (Asia/Tashkent), jarayonning mintaqasi EMAS.

    Sana solishtiruvining YAGONA manbai. `date.today()` to'g'ridan-
    to'g'ri ishlatilmasin — u serverning mintaqasiga bog'liq va
    o'sha bog'liqlik vaqt mintaqasiga bog'liq mo'rt sinovlar
    manbai edi.
    """
    return _dt.datetime.now(BIZNES_TZ).date()


# ---------------------------------------------------------------------------
# KANONIK HUJJAT TURLARI
# ---------------------------------------------------------------------------
# Har tur: kod, o'zbekcha nom, izoh, bazaviymi, aniqlash naqshlari, istisnolar.
#
# `patterns` — variantlar ro'yxati. Har variant O'ZAKLAR KORTEJI: hammasi
#   matnda VA bir-biridan WINDOW belgi ichida bo'lsa — mos keldi.
#   Bitta o'zakli variant faqat so'z o'z-o'zidan aniq bo'lganda ishlatiladi
#   ("доверенност", "ishonchnoma").
# `exclude` — BITTA o'zaklar (ibora emas!): topilgan oyna atrofida shulardan
#   biri bo'lsa moslik BEKOR qilinadi. Ibora yozib bo'lmaydi, chunki rus tili
#   qo'shimchalari orada turadi ("гарантийный срок" != "гарантийн срок").
#
# O'zaklar QISQARTIRILGAN holda yoziladi (so'z oxiri kesilgan): rus tilida
# kelishik qo'shimchalari o'zgaradi — "сертификат / сертификата / сертификатов".
DOC_TYPES: List[Dict[str, Any]] = [
    {
        "code": "reg_certificate",
        "label": "Davlat ro'yxatidan o'tganlik guvohnomasi",
        "hint": "Yuridik shaxsning davlat ro'yxatidan o'tkazilganligi haqidagi guvohnoma.",
        "base": True,
        "patterns": [
            ("свидетельств", "регистрац"),
            ("свидетельств", "государственн"),
            # Apostrof canon() da tushib qoladi -> "ro'yxat" ham, "royxat" ham
            # bir xil qidiriladi. O'zbek KIRILL yozuvi esa alifbo qoidasi bilan
            # chiqmaydi (ў -> у emas), shuning uchun alohida yoziladi.
            ("guvohnoma", "ro'yxat"),
            ("гувохнома", "руйхат"),
            ("гувохнома", "рўйхат"),
            ("davlat ro'yxatidan",),
            ("выписк", "егрюл"),
        ],
        # "система сбора и РЕГИСТРАЦИИ данных" — texnik matn, hujjat emas.
        "exclude": ("датчик", "сигнал"),
    },
    {
        "code": "power_of_attorney",
        "label": "Ishonchnoma",
        "hint": "Ariza va shartnomani imzolovchi vakilga berilgan ishonchnoma.",
        "base": True,
        "patterns": [
            ("доверенност",),
            ("ishonchnoma",),
            ("ишончнома",),
            ("vakolat xat",),
        ],
        "exclude": (),
    },
    {
        "code": "license",
        "label": "Litsenziya / faoliyat ruxsatnomasi",
        "hint": "Litsenziyalanadigan faoliyat turi uchun litsenziya yoki ruxsatnoma.",
        "base": True,
        "patterns": [
            ("лиценз", "деятельност"),
            ("лиценз", "осуществлен"),
            ("лиценз", "вид работ"),
            ("litsenziya", "faoliyat"),
            ("литсензия", "фаолият"),
            ("faoliyat turi uchun litsenziya",),
            ("ruxsatnoma", "faoliyat"),
        ],
        # "Лицензия (бессрочная) для программного обеспечения",
        # "tizimning litsenziyalangan dasturiy ta'minoti" — bu SOTIB
        # OLINAYOTGAN dastur litsenziyasi (tovarning o'zi), yetkazuvchidan
        # talab qilinadigan hujjat EMAS. O'lchangan holatlar: tender 4221601,
        # 20000504422.
        "exclude": ("программн", "software", "бессрочн", "подписк",
                    "dastur", "дастур"),
    },
    {
        "code": "conformity_certificate",
        "label": "Muvofiqlik sertifikati",
        "hint": "Tovarning standart/texnik reglamentga muvofiqligi sertifikati.",
        "base": True,
        "patterns": [
            ("сертификат", "соответстви"),
            ("сертификат", "качеств"),
            ("сертификат", "происхожден"),
            ("sertifikat", "muvofiqlik"),
            ("сертификат", "мувофиклик"),
            ("muvofiqlik sertifikat",),
            ("sifat sertifikat",),
            ("декларац", "соответстви"),
            ("гигиеническ", "сертификат"),
        ],
        # "в соответствии с ..." ("...ga muvofiq") eng ko'p uchraydigan soxta
        # moslik manbai edi: yalang'och "соответстви" 342 tenderдan 58 tasiga
        # mos kelardi. IKKI O'ZAK talabi ("сертификат" ham yonida bo'lsin)
        # ularning hammasini yopdi — alohida istisno kerak emas.
        "exclude": (),
    },
    {
        "code": "guarantee_letter",
        "label": "Kafolat xati",
        "hint": "Yetkazib berish/sifat majburiyatini tasdiqlovchi kafolat xati.",
        "base": True,
        # DIQQAT: XAT (письмо) so'zi SHART. "Гарантийные обязательства",
        # "гарантийный срок эксплуатации" — bular tovarning kafolat SHARTI,
        # topshiriladigan hujjat emas (o'lchangan soxta holat: tender 2115330).
        "patterns": [
            ("гарантийн", "письм"),
            ("письм", "гаранти"),
            ("банковск", "гаранти"),
            ("kafolat xat",),
            ("кафолат хат",),
        ],
        "exclude": ("пробег", "эксплуатац"),
    },
    {
        "code": "bank_details",
        "label": "Bank rekvizitlari",
        "hint": "Hisob raqami, MFO, STIR — to'lov uchun rekvizitlar.",
        "base": True,
        "patterns": [
            ("банковск", "реквизит"),
            ("реквизит", "счет"),
            ("bank rekvizit",),
            ("банк реквизит",),
            ("hisob raqam",),
            ("хисоб раками",),
            ("расчетн счет",),
        ],
        # "банковской системы" (bank tizimi haqidagi qarorlar) anno da ko'p
        # uchraydi — "реквизит" yonida turishi talabi uni allaqachon yopadi.
        "exclude": ("систем", "сектор"),
    },

    # --- Bazaviy emas: faqat tender matnidan aniqlansa qo'shiladi -----------
    {
        "code": "tax_reference",
        "label": "Soliq ma'lumotnomasi (qarzdorlik yo'qligi)",
        "hint": "Soliq organidan qarzdorlik yo'qligi haqidagi ma'lumotnoma.",
        "base": False,
        "patterns": [
            ("справк", "налогов"),
            ("отсутстви", "задолженност"),
            ("налогов", "задолженност"),
            ("soliq", "qarzdorlik"),
            ("солик", "карздорлик"),
            ("soliq malumotnoma",),
        ],
        "exclude": ("аналогов", "каталогов"),
    },
    {
        "code": "charter",
        "label": "Ustav / ta'sis hujjatlari",
        "hint": "Ustav va ta'sis shartnomasining nusxalari.",
        "base": False,
        "patterns": [
            ("устав", "копи"),
            ("учредительн", "документ"),
            ("копи", "устав"),
            ("ustav", "nusxa"),
            ("ta'sis hujjat",),
            ("таъсис хужжат",),
        ],
        "exclude": ("уставн капитал",),
    },
    {
        "code": "financial_report",
        "label": "Moliyaviy hisobot / balans",
        "hint": "Buxgalteriya balansi yoki moliyaviy hisobot nusxasi.",
        "base": False,
        "patterns": [
            ("бухгалтерск", "баланс"),
            ("финансов", "отчетност"),
            ("moliyaviy hisobot",),
            ("buxgalteriya balans",),
            ("бухгалтерия баланс",),
        ],
        # "Разработка ФИНАНСОВОЙ МОДЕЛИ ... Отчет по разработке" — bu xarid
        # qilinayotgan XIZMAT, ariza hujjati emas (tender 3854512, 3943887).
        # Shu sabab ("финансов","отчет") naqshi olib tashlandi: "отчет" juda
        # keng, faqat "отчетност" (hisobot shakli) qoldirildi.
        "exclude": ("модел", "разработк"),
    },
    {
        "code": "technical_proposal",
        "label": "Texnik taklif (texnik topshiriqqa javob)",
        "hint": "Texnik topshiriq talablariga muvofiqlik jadvali/texnik taklif.",
        "base": False,
        "patterns": [
            ("техническ", "предложен"),
            ("техническ задани",),
            ("texnik topshiriq",),
            ("техник топширик",),
            ("texnik taklif",),
        ],
        "exclude": (),
    },
    {
        "code": "price_offer",
        "label": "Narx taklifi",
        "hint": "Tijorat/narx taklifi (smeta bilan).",
        "base": False,
        "patterns": [
            ("ценов", "предложен"),
            ("коммерческ", "предложен"),
            ("narx taklif",),
            ("нарх таклиф",),
        ],
        "exclude": (),
    },
]

#: Kod -> tur (tez topish uchun)
BY_CODE: Dict[str, Dict[str, Any]] = {d["code"]: d for d in DOC_TYPES}

#: Bazaviy (deyarli har tenderда kerak) turlar — biznes-jarayon 10-11 bosqich.
BASE_CODES: List[str] = [d["code"] for d in DOC_TYPES if d["base"]]

#: Cheklist band holatlari
STATUSES = ("ok", "expiring_soon", "expired", "missing")

BASE_EVIDENCE = ("Biznes-jarayonning odatiy ariza to'plami "
                 "(davlat xaridi, 10-11 bosqich)")


# ---------------------------------------------------------------------------
# MATNNI KANONIK SHAKLGA KELTIRISH
# ---------------------------------------------------------------------------
#: O'zbek lotin yozuvidagi apostrof shakllari. Manbada bir xil so'z uch xil
#: yoziladi: "ro'yxat", "ro‘yxat", "roʻyxat". Hammasini olib tashlaymiz —
#: naqshlar ham shu qoidadan o'tadi, ya'ni bitta yozuv yetarli.
_APOSTROPHES = "'‘’ʻ`´"

#: O'ZBEK KIRILL harflari -> eng yaqin ruscha kirill.
#: api/translit.py ruscha kirill uchun sozlangan: uning yig'ish jadvalida
#: ҳ/қ/ў/ғ yo'q, lotin->kirill o'girishi esa "guvohnoma" ni "гувохнома" (х
#: bilan) qiladi. Manbada esa "гувоҳнома" (ҳ bilan) yoziladi va ular
#: uchrashmaydi. Shu qo'shimcha yig'ish ikkalasini bir nuqtaga keltiradi:
#:      гувоҳнома -> гувохнома   <-  guvohnoma
#:      рўйхат    -> руихат      <-  ro'yxat
#: DIQQAT: translit.py O'ZGARTIRILMAYDI — bu faqat shu moduldagi qo'shimcha
#: qatlam (u yerdagi SQL_FOLD bilan mos bo'lishi shart emas: cheklist
#: qidiruvni SQL da emas, Python tomonda bajaradi).
_UZ_CYR_FOLD = {"ҳ": "х", "қ": "к", "ў": "у", "ғ": "г", "ҷ": "ж", "ө": "о"}

_CHAR_CACHE: Dict[str, str] = {}


def _canon_char(ch: str) -> str:
    """Bitta belgining kanonik shakli (bo'sh satr = tushib qoladi).

    translit.norm_text() belgi-ba-belgi ishlaydi (kichik harf + yig'ish,
    ь/ъ o'chadi), shuning uchun butun matnni belgi bo'yicha o'tkazish
    norm_text(matn) bilan bir xil natija beradi. Bizga belgi bo'yicha kerak —
    dalil (evidence) uchun XOM matndagi pozitsiya saqlanishi shart.
    """
    v = _CHAR_CACHE.get(ch)
    if v is None:
        if ch in _APOSTROPHES:
            v = ""
        else:
            low = ch.lower()
            v = translit.norm_text(_UZ_CYR_FOLD.get(low, low))
        _CHAR_CACHE[ch] = v
    return v


def canon(s: str) -> str:
    """Matnning kanonik shakli (naqsh va matn uchun bir xil quvur)."""
    return "".join(_canon_char(c) for c in s or "")


def _canon_indexed(raw: str) -> Tuple[str, List[int]]:
    """Kanonik matn + har kanonik belgining XOM matndagi indeksi.

    Yig'ish ba'zi belgilarni o'chiradi (ь, ъ, apostrof), shuning uchun
    pozitsiyalar siljiydi. Xarita busiz dalil noto'g'ri joydan kesilardi.
    """
    parts: List[str] = []
    idx: List[int] = []
    for i, ch in enumerate(raw):
        f = _canon_char(ch)
        if not f:
            continue
        parts.append(f)
        idx.extend([i] * len(f))
    return "".join(parts), idx


# ---------------------------------------------------------------------------
# NAQSH ANIQLASH
# ---------------------------------------------------------------------------
@lru_cache(maxsize=512)
def _stem_variants(stem: str) -> Tuple[str, ...]:
    """O'zakning barcha alifbo yozuvlari — KANONIK shaklda.

    "сертификат" -> kirill yig'ilgan + "sertifikat" + boshqa o'qilishlar.
    Ya'ni o'zakni bitta alifboda yozish yetarli (api/translit.py variants()).
    Alifbo bilan bog'lanmagan so'zlar (свидетельство <-> guvohnoma) esa
    naqshlarda ALOHIDA yoziladi — translit tarjimon emas.
    """
    base = canon(stem)
    out, seen = [], set()
    for v in translit.variants(stem):
        c = canon(v)
        if not c or c in seen:
            continue
        # QISQARIB ketgan variantni tashlaymiz. Misol: "счет" -> lotinga
        # "schet" -> qayta kirillga "шет" (sch -> ш). Uch harfli "шет"
        # "реШЕТка" ga ham mos kelib, soxta natija berardi. Faqat o'zakning
        # to'liq uzunligidagi variant bunday qisqa bo'lishi mumkin.
        if len(c) < 4 and len(c) < len(base):
            continue
        seen.add(c)
        out.append(c)
    return tuple(out)


def _find_all(hay: str, needles: Sequence[str]) -> List[Tuple[int, int]]:
    """`needles` dan har birining barcha uchrashlari: (boshlanish, tugash)."""
    out: List[Tuple[int, int]] = []
    for n in needles:
        start = 0
        while True:
            i = hay.find(n, start)
            if i < 0:
                break
            out.append((i, i + len(n)))
            start = i + 1
    return sorted(out)


def _match_alternative(blob: str, stems: Sequence[str],
                       exclude: Sequence[str]) -> Optional[Tuple[int, int]]:
    """Bitta variant (o'zaklar korteji) matnда bormi?

    Hamma o'zak topilishi VA ular WINDOW belgi ichida bo'lishi shart.
    Topilgan oyna atrofida `exclude` o'zaklaridan biri bo'lsa — moslik
    BEKOR qilinadi (o'lchangan soxta natijalarni yopadi).
    Natija: kanonik matndagi (boshlanish, tugash) — dalil kesish uchun.
    """
    per_stem: List[List[Tuple[int, int]]] = []
    for st in stems:
        pos = _find_all(blob, _stem_variants(st))
        if not pos:
            return None
        per_stem.append(pos)

    # Birinchi o'zakning har uchrashi uchun qolganlari yaqinda turibdimi?
    for a0, b0 in per_stem[0]:
        lo, hi, ok = a0, b0, True
        for pos in per_stem[1:]:
            near = [p for p in pos if abs(p[0] - a0) <= WINDOW]
            if not near:
                ok = False
                break
            a, b = min(near, key=lambda p: abs(p[0] - a0))
            lo, hi = min(lo, a), max(hi, b)
        if not ok:
            continue
        ctx = blob[max(0, lo - 40):hi + 40]
        if any(v in ctx for ex in exclude for v in _stem_variants(ex)):
            continue
        return lo, hi
    return None


def _evidence(raw: str, idx: Sequence[int], lo: int, hi: int) -> str:
    """Dalil bo'lagi — foydalanuvchi yig'ilgan emas, ASL matnni ko'radi."""
    if not idx:
        return ""
    a = idx[max(0, lo - EVIDENCE_PAD)]
    b = idx[min(len(idx) - 1, hi + EVIDENCE_PAD)] + 1
    frag = re.sub(r"\s+", " ", raw[a:b]).strip()
    return ("…" if a > 0 else "") + frag + ("…" if b < len(raw) else "")


def detect_required(tender_texts: Sequence[Any]) -> List[Dict[str, Any]]:
    """Tender matnidan majburiy hujjatlarni aniqlaydi.

    `tender_texts` — matn manbalari. Har element:
        {"source": "Tender izohi", "text": "..."}  yoki  ("Tender izohi", "...")
        yoki oddiy satr (manba nomi "tender matni" bo'ladi).

    Natija: [{doc_type, label, evidence, source, confidence}] — faqat
    TOPILGANLARI. Bazaviy ro'yxat bu yerga QO'SHILMAYDI (uni check() qo'shadi),
    chunki bu funksiyaning vazifasi — "tender AYNAN nimani so'ragan".
    """
    found: Dict[str, Dict[str, Any]] = {}

    for item in tender_texts or []:
        if isinstance(item, dict):
            source, raw = item.get("source") or "tender matni", item.get("text")
        elif isinstance(item, (tuple, list)) and len(item) == 2:
            source, raw = item
        else:
            source, raw = "tender matni", item
        if not raw:
            continue
        raw = str(raw)
        blob, idx = _canon_indexed(raw)

        for d in DOC_TYPES:
            if d["code"] in found:
                continue  # birinchi (eng ishonchli) dalil yetarli
            for stems in d["patterns"]:
                hit = _match_alternative(blob, stems, d.get("exclude") or ())
                if not hit:
                    continue
                lo, hi = hit
                found[d["code"]] = {
                    "doc_type": d["code"],
                    "label": d["label"],
                    "source": source,
                    "evidence": _evidence(raw, idx, lo, hi),
                    # Bir o'zakli qoida kuchsizroq dalil: so'z boshqa ma'noda
                    # kelgan bo'lishi mumkin. Ko'p o'zakli qoida ancha aniq.
                    "confidence": 90 if len(stems) > 1 else 70,
                }
                break

    # DOC_TYPES tartibini saqlaymiz (interfeys tartibi barqaror bo'lsin)
    return [found[d["code"]] for d in DOC_TYPES if d["code"] in found]


# ---------------------------------------------------------------------------
# KOMPANIYA BAZASI bilan solishtirish
# ---------------------------------------------------------------------------
def _as_date(v: Any) -> Optional[_dt.date]:
    """Har qanday ko'rinishdagi qiymatni BIZNES SANASIGA aylantiradi.

    MINTAQALI `datetime` BIZNES MINTAQASIGA o'giriladi, keyin sanasi
    olinadi. Aks holda UTC da saqlangan `2026-08-30T22:00Z` "30-avgust"
    bo'lardi, holbuki Toshkentda u allaqachon 31-avgust soat 03:00.
    Hozir `valid_until` — `date` ustuni, ya'ni bu yo'l ishlamaydi,
    lekin funksiya umumiy va keyin `timestamptz` bilan ham
    chaqirilishi mumkin.
    """
    if v is None or v == "":
        return None
    if isinstance(v, _dt.datetime):
        if v.tzinfo is not None:
            v = v.astimezone(BIZNES_TZ)
        return v.date()
    if isinstance(v, _dt.date):
        return v
    try:
        return _dt.date.fromisoformat(str(v)[:10])
    except ValueError:
        return None


def doc_status(doc: Optional[Dict[str, Any]],
               today: Optional[_dt.date] = None) -> str:
    """Bitta hujjatning holati: missing | expired | expiring_soon | ok.

    ══════════════════ QOIDA (yagona manba) ══════════════════

    BIZNES KUNI — Asia/Tashkent (UTC+5, yozgi vaqt yo'q), `bugun()`.

    `valid_until` — hujjat yaroqli bo'lgan OXIRGI KUN va u KIRADI:
    o'sha kunning oxirigacha hujjat AMAL QILADI. Ya'ni "bugun
    tugaydi" degan hujjat BUGUN hali yaroqli.

        hujjat yo'q                              -> missing
        valid_until IS NULL                      -> ok
              (MUDDATSIZ, "noma'lum" EMAS)
        valid_until <  bugun                     -> expired
        0 <= (valid_until - bugun) <= CHEGARA     -> expiring_soon
        (valid_until - bugun) >  CHEGARA          -> ok

    USTUVORLIK: `expired` `expiring_soon` DAN OLDIN tekshiriladi —
    muddati o'tgan hujjat "tugayapti" bo'lib ko'rinmasin.

    CHEGARA IKKI TOMONDAN KIRADI (`EXPIRING_SOON_DAYS` = 30):
        bugun            -> expiring_soon   (0 kun qoldi, hali yaroqli)
        bugun + 1        -> expiring_soon
        bugun + 30       -> expiring_soon   (aynan chegarada)
        bugun + 31       -> ok
        kecha            -> expired
    """
    if not doc:
        return "missing"
    today = today or bugun()
    vu = _as_date(doc.get("valid_until"))
    if vu is None:
        return "ok"          # muddatsiz — "ma'lumot yo'q" emas, "cheklanmagan"
    if vu < today:
        return "expired"
    if (vu - today).days <= EXPIRING_SOON_DAYS:
        return "expiring_soon"
    return "ok"


_STATUS_RANK = {"ok": 0, "expiring_soon": 1, "expired": 2, "missing": 3}


def _pick_best(docs: Sequence[Dict[str, Any]],
               today: _dt.date) -> Optional[Dict[str, Any]]:
    """Bir turdagi bir necha hujjatdan ENG YAROQLISINI tanlaydi.

    Kompaniyada eski va yangilangan nusxa birga turishi mumkin — eskisi
    tufayli butun band "muddati tugagan" bo'lib qolmasin.
    """
    if not docs:
        return None
    def key(d):
        vu = _as_date(d.get("valid_until"))
        # muddatsiz eng yaxshi (0), keyin uzoq muddatlisi
        return (_STATUS_RANK[doc_status(d, today)], 0 if vu is None else 1,
                -(vu.toordinal() if vu else 0))
    return sorted(docs, key=key)[0]


def shape_document(r: Dict[str, Any],
                   today: Optional[_dt.date] = None) -> Dict[str, Any]:
    """DB qatorini JSON ga tayyorlaydi (endpoint ham, cheklist ham ishlatadi)."""
    def iso(v):
        d = _as_date(v)
        return d.isoformat() if d else None
    return {
        "id": r.get("id"),
        "doc_type": r.get("doc_type"),
        "label": (BY_CODE.get(r.get("doc_type")) or {}).get("label"),
        "name": r.get("name"),
        "number": r.get("number"),
        "issued_at": iso(r.get("issued_at")),
        "valid_until": iso(r.get("valid_until")),
        "file_name": r.get("file_name"),
        # ESKI MATN HAVOLASI — yangi qatorlarda `null`. Qoldirilgan,
        # chunki mavjud 13 qatorda u yagona ma'lumot.
        "file_ref": r.get("file_ref"),
        # HAQIQIY FAYL. `str()` SHART: `uuid.UUID` JSON ga
        # serializatsiya BO'LMAYDI va endpoint 500 berardi.
        "yuklama_id": (str(r["yuklama_id"])
                       if r.get("yuklama_id") else None),
        "note": r.get("note"),
        "status": doc_status(r, today),
        "days_left": _days_left(r.get("valid_until"), today),
    }


def _days_left(valid_until: Any,
               today: Optional[_dt.date] = None) -> Optional[int]:
    """Muddatgacha qolgan kunlar. `today` — SINOV uchun: `doc_status()` va
    `build_checklist()` sanani parametr sifatida oladi, bu funksiya esa
    avval har doim haqiqiy bugundan hisoblardi. Natijada bitta javob ichida
    `status` bir sanaga, `days_left` boshqasiga qarab hisoblanib, sinov
    fixture'lari eskirgach ular bir-biriga zid bo'lib qolardi."""
    vu = _as_date(valid_until)
    if vu is None:
        return None
    return (vu - (today or bugun())).days


def build_checklist(detected: Sequence[Dict[str, Any]],
                    company_docs: Sequence[Dict[str, Any]],
                    today: Optional[_dt.date] = None) -> Dict[str, Any]:
    """Cheklistni yig'adi — BAZAVIY ro'yxat + tenderda topilganlari.

    DB'ga bog'liq emas: sof funksiya, shuning uchun sinovi oson.
    """
    today = today or bugun()
    det = {d["doc_type"]: d for d in detected or []}

    # Turlar bo'yicha guruhlash (notanish kodlar ham saqlanadi)
    by_type: Dict[str, List[Dict[str, Any]]] = {}
    for r in company_docs or []:
        by_type.setdefault(r.get("doc_type"), []).append(r)

    codes = [c for c in BASE_CODES]
    for c in [d["code"] for d in DOC_TYPES]:
        if c in det and c not in codes:
            codes.append(c)

    items: List[Dict[str, Any]] = []
    for code in codes:
        meta = BY_CODE[code]
        hit = det.get(code)
        best = _pick_best(by_type.get(code) or [], today)
        st = doc_status(best, today)
        items.append({
            "doc_type": code,
            "label": meta["label"],
            "hint": meta["hint"],
            # TENDERда topilgani bazaviydan kuchliroq: dalil bor.
            "required_by": "tender" if hit else "bazaviy",
            "evidence": hit["evidence"] if hit else BASE_EVIDENCE,
            "evidence_source": hit["source"] if hit else None,
            "confidence": hit["confidence"] if hit else None,
            "in_base": best is not None,
            "document": shape_document(best, today) if best else None,
            "status": st,
            "days_left": (_days_left(best.get("valid_until"), today)
                          if best else None),
        })

    # Kompaniyada bor, lekin cheklistга kirmagan hujjatlar — ularni ham
    # ko'rsatamiz (broker "bu ham kerak bo'lishi mumkin" deb ko'radi),
    # lekin ALOHIDA ro'yxatда, majburiy sifatida emas.
    listed = {i["doc_type"] for i in items}
    extra = []
    for code, rows in by_type.items():
        if code in listed:
            continue
        best = _pick_best(rows, today)
        extra.append({
            "doc_type": code,
            "label": (BY_CODE.get(code) or {}).get("label") or code,
            "document": shape_document(best, today) if best else None,
            "status": doc_status(best, today),
        })
    extra.sort(key=lambda x: x["label"])

    n_missing = sum(1 for i in items if i["status"] == "missing")
    n_expired = sum(1 for i in items if i["status"] == "expired")
    n_soon = sum(1 for i in items if i["status"] == "expiring_soon")
    n_ok = sum(1 for i in items if i["status"] == "ok")

    detected_any = bool(det)
    if detected_any:
        note = (f"Tender matnidan {len(det)} ta hujjat talabi aniqlandi. "
                "Qolganlari — odatiy ariza to'plami.")
    else:
        note = ("Tender matnida aniq hujjat talabi topilmadi — bazaviy ro'yxat "
                "ko'rsatilmoqda. To'liq talablar tenderning biriktirilgan "
                "hujjatlarida bo'lishi mumkin, ularni qo'lda tekshiring.")

    return {
        "items": items,
        "extra_documents": extra,
        "summary": {
            "total": len(items),
            "ready": n_ok,
            "missing": n_missing,
            "expired": n_expired,
            "expiring_soon": n_soon,
            # Ariza berishga to'sqinlik qiladiganlar (yo'q + muddati tugagan)
            "blocking": n_missing + n_expired,
            "detected_from_tender": detected_any,
            "detected_count": len(det),
            "note": note,
            # Cheklist nima QILMAYDI — noto'g'ri kafolat hissini yo'qotamiz.
            "disclaimer": ("Cheklist hujjat BORLIGINI va MUDDATINI tekshiradi, "
                           "mazmunining huquqiy to'g'riligini emas."),
        },
    }


# ---------------------------------------------------------------------------
# DB QATLAMI — endpoint shu ikki funksiyani chaqiradi
# ---------------------------------------------------------------------------
#: Tender matn manbalari. `tender_document_text` (boshqa modul jadvali) —
#: IXTIYORIY: mavjud bo'lsa qo'shiladi, bo'lmasa so'rov shunchaki NULL beradi.
#: to_regclass() bilan tekshiramiz — jadval yo'q bo'lsa xato chiqmasin.
TENDER_TEXTS_SQL = """
SELECT t.name          AS tender_name,
       d.anno          AS anno,
       d.method_marks  AS method_marks,
       d.offer_period  AS offer_period,
       (SELECT string_agg(DISTINCT COALESCE(i.spec, ''), E'\\n')
          FROM tender_item i WHERE i.tender_id = t.id)   AS item_spec,
       (SELECT string_agg(DISTINCT COALESCE(l.title, ''), E'\\n')
          FROM tender_lot l WHERE l.tender_id = t.id)    AS lot_titles
FROM tender t
LEFT JOIN tender_detail d ON d.tender_id = t.id
WHERE t.id = %(tender_id)s
"""

#: Biriktirilgan hujjatlardan ajratilgan matn (agar boshqa modul to'ldirgan
#: bo'lsa). Jadval yo'q bo'lishi MUMKIN — chaqiruvchi xatoni yutadi.
DOC_TEXT_SQL = """
SELECT string_agg(x.text, E'\\n') AS doc_text
FROM (SELECT text FROM tender_document_text
      WHERE tender_id = %(tender_id)s AND text IS NOT NULL
      LIMIT 20) x
"""

_TEXT_LABELS = {
    "tender_name": "Tender nomi",
    "anno": "Tender izohi (anno)",
    "method_marks": "Baholash usuli",
    "offer_period": "Takliflar muddati",
    "item_spec": "Pozitsiya spetsifikatsiyasi",
    "lot_titles": "Lot nomlari",
    "doc_text": "Biriktirilgan hujjat matni",
}

# `yuklama_id` RO'YXATDA ham kerak: interfeys "yuklab olish"
# tugmasini FAQAT shu maydonga qarab chizadi. Usiz fayl yuklangan
# bo'lsa ham tugma ko'rinmasdi va foydalanuvchi faylni QAYTA
# yuklashga urinardi.
_DOC_COLS = ("id, doc_type, name, number, issued_at, valid_until, "
             "file_name, file_ref, yuklama_id, note, created_at, updated_at")

# KOMPANIYA HUJJATLARI — ijarachi siri (J1.6). Har so'rovda `company_id`.
# `id` bo'yicha murojaatda ham filtr bor: begona hujjatni taxmin qilib
# tahrirlash/o'chirish mumkin bo'lmasin (IDOR) — javob 404 bo'ladi.
DOCS_LIST_SQL = (f"SELECT {_DOC_COLS} FROM company_document "
                 "WHERE company_id = %(company_id)s ORDER BY doc_type, id")

DOC_INSERT_SQL = f"""
INSERT INTO company_document
    (company_id, doc_type, name, number, issued_at, valid_until,
     file_name, file_ref, note)
VALUES (%(company_id)s, %(doc_type)s, %(name)s, %(number)s, %(issued_at)s,
        %(valid_until)s, %(file_name)s, %(file_ref)s, %(note)s)
RETURNING {_DOC_COLS}
"""

DOC_UPDATE_SQL = f"""
UPDATE company_document SET
    doc_type=%(doc_type)s, name=%(name)s, number=%(number)s,
    issued_at=%(issued_at)s, valid_until=%(valid_until)s,
    file_name=%(file_name)s, file_ref=%(file_ref)s, note=%(note)s,
    updated_at=now()
WHERE id=%(id)s AND company_id=%(company_id)s
RETURNING {_DOC_COLS}
"""

DOC_DELETE_SQL = ("DELETE FROM company_document "
                  "WHERE id=%(id)s AND company_id=%(company_id)s RETURNING id")


def tender_texts(tender_id: int) -> List[Dict[str, str]]:
    """Tenderning barcha matn manbalari (manba nomi bilan — dalil uchun)."""
    from api import db  # kech import: modulni DB'siz ham sinash mumkin bo'lsin

    row = db.query_one(TENDER_TEXTS_SQL, {"tender_id": tender_id})
    if not row:
        return []
    out = [{"source": _TEXT_LABELS[k], "text": v}
           for k, v in row.items() if v and str(v).strip()]

    # Hujjat matni — ixtiyoriy manba (jadval boshqa modulniki).
    try:
        r2 = db.query_one(DOC_TEXT_SQL, {"tender_id": tender_id})
        if r2 and r2.get("doc_text"):
            out.append({"source": _TEXT_LABELS["doc_text"], "text": r2["doc_text"]})
    except Exception:
        pass  # jadval hali yo'q yoki bo'sh — cheklist busiz ham ishlaydi
    return out


def check(tender_id: int,
          docs: Optional[Sequence[Dict[str, Any]]] = None,
          company_id: Optional[int] = None) -> Dict[str, Any]:
    """Tender bo'yicha to'liq cheklist — endpoint shuni qaytaradi.

    `docs` — hujjatlar MANBASI. Berilmasa (odatiy hol) BROKER kompaniyasining
    hujjatlari (`company_document`) olinadi — eski xatti-harakat o'zgarmaydi.

    ERP 2-bosqichi buni MIJOZ nomidan qatnashish uchun ishlatadi: opportunity
    kartasi o'z mijozining hujjatlarini uzatadi va cheklist o'shalarga qarab
    hisoblanadi. Qoidalar (`detect_required`, `build_checklist`) o'zgarmaydi —
    ular allaqachon manbadan mustaqil, sof funksiyalar.

    NEGA `docs`, `client_id` EMAS: ERP — ALOHIDA loyiha va o'z bazasidagi
    mijoz hujjatlarini o'zi biladi. `client_id` qabul qilinsa shu modul erp
    sxemasidan o'qishi kerak bo'lardi; `docs` bilan esa u hech qanday tashqi
    tizimni bilmaydi — kirish oddiy ro'yxat, qoidalar shu yerda qoladi.
    """
    from api import db

    texts = tender_texts(tender_id)
    if docs is None:
        # J1.6: hujjatlar SHU kompaniyaniki. `company_id` berilmasa —
        # sessiyasiz chaqiruv (sinov, ERP): yagona faol hisob olinadi.
        if company_id is None:
            from api import auth
            company_id = auth.sole_company_id()
        docs = db.query(DOCS_LIST_SQL, {"company_id": company_id})
    res = build_checklist(detect_required(texts), docs)
    res["tender_id"] = tender_id
    res["text_sources"] = [t["source"] for t in texts]
    return res


# ---------------------------------------------------------------------------
# HUJJATLAR SHABLONI — yuklab olish va to'ldirilgan holda qaytarib yuklash
# ===========================================================================
# NEGA SHABLON FAYL, ekranda ro'yxat emas: hujjatlar bazasini birinchi marta
# to'ldirish — 11 ta formani qo'lda kiritish demakdir. Shablon esa TALAB
# ETILADIGAN HUJJATLAR RO'YXATI bilan OLDINDAN TO'LDIRILGAN holda keladi:
# broker faqat raqam va sanalarni yozadi. Sinov ma'lumotini kiritish ham
# shu yo'l bilan bir marta bajariladi.
#
# Ro'yxat DOC_TYPES dan olinadi — cheklist AYNAN shu turlarni tekshiradi,
# ya'ni shablonni to'ldirish = cheklistni yopish. Ikkinchi (qo'lda yozilgan)
# ro'yxat bo'lsa, ular vaqt o'tib bir-biridan ajralib ketardi.
#
# FORMAT: .xlsx / .csv — `api/importer.py` bilan bir xil quvur (o'sha
# `read_table`, `norm_header`, `cell_text`). Katalog importi bilan bir xil
# tajriba: dry-run -> ko'rish -> tasdiqlash.
# ---------------------------------------------------------------------------
#: Shablon ustunlari: (maydon, sarlavha, tanish uchun aliaslar)
#: Ustunlar TARTIBI muhim emas — sarlavha bo'yicha tanilaadi (katalog importi
#: bilan bir xil qoida), lekin fayldagi tartib shu ro'yxatga teng.
TEMPLATE_COLUMNS: List[Tuple[str, str, Tuple[str, ...]]] = [
    ("doc_type", "Hujjat turi", (
        "hujjat turi", "tur", "turi", "tip", "тип документа", "тип",
        "вид документа", "doc type", "type")),
    ("name", "Hujjat nomi", (
        "hujjat nomi", "nomi", "nom", "nomlanishi", "наименование",
        "название документа", "название", "name", "document name")),
    ("number", "Raqami", (
        "raqami", "raqam", "seriya raqami", "номер", "№", "number", "no")),
    ("issued_at", "Berilgan sana", (
        "berilgan sana", "berilgan", "berilgan sanasi", "дата выдачи",
        "выдан", "issued", "issue date", "issued at")),
    ("valid_until", "Amal qiladi (gacha)", (
        "amal qiladi", "amal qilish muddati", "muddati", "amal qiladi gacha",
        "срок действия", "действует до", "годен до", "valid until",
        "expiry", "expires")),
    ("file_name", "Fayl nomi", (
        "fayl nomi", "fayl", "имя файла", "файл", "file", "file name")),
    ("file_ref", "Havola / yo‘l", (
        "havola", "havola yo'l", "yo'l", "manzil", "ссылка", "путь",
        "link", "url", "path", "file ref")),
    ("note", "Izoh", (
        "izoh", "izohlar", "примечание", "комментарий", "note", "comment")),
]

TEMPLATE_HEADERS: List[str] = [h for _, h, _ in TEMPLATE_COLUMNS]

#: Sarlavha -> maydon. Eng uzun alias birinchi (katalog importidagi qoida:
#: "amal qilish muddati" "muddati" dan ustun bo'lsin).
_COL_ALIASES: List[Tuple[str, str]] = []

#: "Amal qiladi" ustunidagi MUDDATSIZ so'zlari. Bo'sh katak ham muddatsiz
#: degani (`valid_until = NULL`), lekin so'z bilan yozish aniqroq.
PERPETUAL_WORDS = ("muddatsiz", "мудатсиз", "муддатсиз", "бессрочно",
                   "бессрочный", "бессрочная", "cheklanmagan", "doimiy",
                   "perpetual", "unlimited", "no expiry")

#: Sana formatlari — Excel'dan matn sifatida kelganda.
_DATE_FORMATS = ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y",
                 "%Y/%m/%d", "%d.%m.%y")


def _build_col_aliases() -> None:
    pairs = []
    for field, header, aliases in TEMPLATE_COLUMNS:
        for a in (header, field) + aliases:
            pairs.append((_importer().norm_header(a), field))
    _COL_ALIASES[:] = sorted(set(pairs), key=lambda x: -len(x[0]))


def _importer():
    from api import importer  # kech import: compliance DB/openpyxl'siz yuklansin
    return importer


def _match_column(header: Any) -> Optional[str]:
    """Sarlavha -> maydon nomi (aniq moslik, keyin ichiga kirish)."""
    if not _COL_ALIASES:
        _build_col_aliases()
    h = _importer().norm_header(header)
    if not h:
        return None
    for alias, field in _COL_ALIASES:
        if h == alias:
            return field
    for alias, field in _COL_ALIASES:
        if len(alias) >= 4 and alias in h:
            return field
    return None


# --- "Hujjat turi" katagini kanonik kodga keltirish -------------------------
#: canon(matn) -> doc_type kodi. Kod ("license"), o'zbekcha nom va uning
#: qisqartmalari qabul qilinadi.
#:
#: Har kalit ALIFBO VARIANTLARI bilan indekslanadi (_stem_variants —
#: aniqlash naqshlaridagi o'sha quvur): shablon Excel'da tahrirlanganda
#: broker turni o'z yozuvida qayta yozishi mumkin — "Литсензия" (o'zbek
#: kirill), "Litsenziya" (lotin), "Лицензия" (rus) uchalasi bir kodga
#: tushishi kerak. Alifbo bilan bog'lanmagan atamalar (guvohnoma <->
#: свидетельство) esa quyida ALOHIDA yoziladi — translit tarjimon emas.
_TYPE_INDEX: Dict[str, str] = {}


def _index_type(key: str, code: str) -> None:
    for v in (canon(key),) + _stem_variants(key):
        if v:
            _TYPE_INDEX.setdefault(v, code)


def _build_type_index() -> None:
    for d in DOC_TYPES:
        for key in (d["code"], d["label"]):
            _index_type(key, d["code"])
    # Qo'lda yoziladigan qisqa nomlar — broker to'liq nomni ko'chirmasligi mumkin
    for alias, code in (
        ("guvohnoma", "reg_certificate"),
        ("davlat royxatidan otganlik guvohnomasi", "reg_certificate"),
        ("свидетельство о регистрации", "reg_certificate"),
        ("ishonchnoma", "power_of_attorney"),
        ("доверенность", "power_of_attorney"),
        ("litsenziya", "license"),
        ("лицензия", "license"),
        ("sertifikat", "conformity_certificate"),
        ("muvofiqlik sertifikati", "conformity_certificate"),
        ("сертификат соответствия", "conformity_certificate"),
        ("kafolat xati", "guarantee_letter"),
        ("гарантийное письмо", "guarantee_letter"),
        ("bank rekvizitlari", "bank_details"),
        ("банковские реквизиты", "bank_details"),
        ("soliq malumotnomasi", "tax_reference"),
        ("налоговая справка", "tax_reference"),
        ("ustav", "charter"),
        ("устав", "charter"),
        ("moliyaviy hisobot", "financial_report"),
        ("финансовая отчетность", "financial_report"),
        ("texnik taklif", "technical_proposal"),
        ("техническое предложение", "technical_proposal"),
        ("narx taklifi", "price_offer"),
        ("ценовое предложение", "price_offer"),
    ):
        _index_type(alias, code)


def match_doc_type(raw: Any) -> Optional[str]:
    """"Muvofiqlik sertifikati" / "conformity_certificate" -> kod.

    Topilmasa None — qator xatoga tushadi. TAXMIN QILINMAYDI: noto'g'ri
    turga tushgan hujjat cheklistda ko'rinmay qolardi, bu esa jimgina
    yo'qotish bo'lardi.
    """
    if not _TYPE_INDEX:
        _build_type_index()
    text = str(raw or "").strip()
    if not text:
        return None

    forms = [f for f in ((canon(text),) + _stem_variants(text)) if f]
    for f in forms:
        if f in _TYPE_INDEX:
            return _TYPE_INDEX[f]
    # Qavs ichidagi izoh yoki ortiqcha so'z bo'lsa — ichiga kirish bo'yicha.
    # Eng UZUN kalit g'olib: "Litsenziya" "Muvofiqlik sertifikati" ichida
    # yo'q, lekin qisqa kalitlar bir-birining ichiga tushib qolmasin.
    for key in sorted(_TYPE_INDEX, key=len, reverse=True):
        if len(key) >= 8 and any(key in f for f in forms):
            return _TYPE_INDEX[key]
    return None


def parse_date(raw: Any) -> Tuple[Optional[_dt.date], Optional[str], bool]:
    """Katakdan sana. Qaytaradi: (sana, xato, muddatsizmi).

    Bo'sh katak -> (None, None, False): "ko'rsatilmagan". Bu MUDDATSIZ bilan
    bir xil emas — chaqiruvchi ularni ajratadi (bo'sh "Amal qiladi" ustuni
    ogohlantirish beradi, chunki cheklist NULL ni "cheklanmagan" deb o'qiydi).
    """
    if raw is None:
        return None, None, False
    if isinstance(raw, _dt.datetime):
        return raw.date(), None, False
    if isinstance(raw, _dt.date):
        return raw, None, False

    s = re.sub(r"\s+", " ", str(raw)).strip()
    if not s:
        return None, None, False
    if canon(s) in {canon(w) for w in PERPETUAL_WORDS}:
        return None, None, True
    for fmt in _DATE_FORMATS:
        try:
            return _dt.datetime.strptime(s, fmt).date(), None, False
        except ValueError:
            continue
    # Excel ba'zan sanani seriya raqami sifatida beradi (1899-12-30 dan kunlar)
    if re.fullmatch(r"\d{5}", s):
        try:
            return (_dt.date(1899, 12, 30) + _dt.timedelta(days=int(s))), None, False
        except (ValueError, OverflowError):
            pass
    return None, (f"sana o‘qilmadi: “{s}” — 31.12.2026 yoki 2026-12-31 "
                  f"ko‘rinishida yozing (yoki “muddatsiz”)"), False


# --- Shablon fayllarini yasash ---------------------------------------------
#: Shablonning BIRINCHI qatori — TO'LDIRILGAN misol. Qolgan qatorlar
#: DOC_TYPES dan avtomatik yasaladi, ya'ni shablon ro'yxat o'zgarsa
#: o'z-o'zidan yangilanadi.
_EXAMPLE_ROW = ["Davlat ro‘yxatidan o‘tganlik guvohnomasi",
                "Davlat ro‘yxatidan o‘tganlik guvohnomasi",
                "AA 1234567", "12.03.2019", "muddatsiz",
                "guvohnoma.pdf", "https://disk.example/guvohnoma.pdf",
                "misol qator — o‘chirib, o‘zingiznikini yozing"]

TEMPLATE_HELP = [
    "SHABLONNI QANDAY TO‘LDIRASIZ",
    "",
    "1. “Hujjat turi” ustuniga TEGMANG — u oldindan to‘ldirilgan. Bu ro‘yxat",
    "   tender cheklisti tekshiradigan hujjatlarning to‘liq ro‘yxati.",
    "2. O‘zingizda BOR hujjatlar qatorini to‘ldiring: raqami, sanalari, havolasi.",
    "3. Hujjat YO‘Q bo‘lsa — qatorni BO‘SH qoldiring. Bo‘sh qator import",
    "   qilinmaydi va cheklistda “yo‘q” bo‘lib turaveradi.",
    "4. Bitta turda bir nechta hujjat bo‘lsa (masalan 2 ta litsenziya) —",
    "   qatorni nusxalab, “Hujjat nomi” ni har xil yozing.",
    "",
    "USTUNLAR",
    "   Hujjat turi         — o‘zgartirmang (kanonik ro‘yxat).",
    "   Hujjat nomi         — hujjatning haqiqiy nomi. MAJBURIY.",
    "   Raqami              — seriya/raqam, masalan “AA 1234567”.",
    "   Berilgan sana       — 31.12.2026 yoki 2026-12-31.",
    "   Amal qiladi (gacha) — muddat tugash sanasi yoki “muddatsiz”.",
    "   Fayl nomi / Havola  — MVP da fayl saqlanmaydi, faqat nom va havola.",
    "   Izoh                — ixtiyoriy.",
    "",
    "DIQQAT — “Amal qiladi” ustuni eng muhimi:",
    "   Cheklist hujjat BOR-YO‘QLIGINI va MUDDATINI tekshiradi. Bo‘sh",
    "   qoldirilsa muddat CHEKLANMAGAN deb o‘qiladi va hujjat doim yaroqli",
    "   ko‘rinadi. Muddati bo‘lsa — albatta yozing.",
    "",
    "QAYTA YUKLASH",
    "   Bir xil “Hujjat turi” + “Hujjat nomi” juftligi bazada topilsa qator",
    "   YANGILANADI, topilmasa QO‘SHILADI. Ya'ni shablonni to‘ldirib qayta",
    "   yuklash xavfsiz — nusxa ko‘paymaydi.",
    "",
    "FORMAT",
    "   .xlsx yoki .csv. Google Sheets: Fayl > Yuklab olish > CSV.",
]


def _template_rows() -> List[List[Any]]:
    """Shablon qatorlari: misol qator + har kanonik tur uchun bo'sh qator."""
    rows: List[List[Any]] = [list(_EXAMPLE_ROW)]
    for d in DOC_TYPES:
        rows.append([d["label"], d["label"], "", "", "", "", "",
                     "Bazaviy — deyarli har tenderda so‘raladi" if d["base"]
                     else "Tenderga qarab so‘raladi"])
    return rows


def template_xlsx() -> bytes:
    """Namunaviy .xlsx shablon: 1-varaq — hujjatlar, 2-varaq — yo'riqnoma."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Hujjatlar"
    ws.append(TEMPLATE_HEADERS)
    head_fill = PatternFill("solid", fgColor="E8EEF7")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = head_fill
        c.alignment = Alignment(vertical="center")

    rows = _template_rows()
    for r in rows:
        ws.append(r)

    # Misol qator — kulrang kursiv: to'ldirilishi kerak bo'lgan qatorlardan
    # ko'rinib tursin (foydalanuvchi uni o'chiradi).
    for c in ws[2]:
        c.font = Font(italic=True, color="8A94A6")
    # Bazaviy turlar — quyuqroq fon: qaysi biri deyarli har doim kerakligi
    # faylning O'ZIDA ko'rinsin (ekranga qaytish shart bo'lmasin).
    base_fill = PatternFill("solid", fgColor="F3F7FF")
    for i, d in enumerate(DOC_TYPES, start=3):   # 1-sarlavha, 2-misol
        if d["base"]:
            for c in ws[i]:
                c.fill = base_fill
        ws.cell(row=i, column=1).font = Font(bold=True)

    for i, w in enumerate([38, 38, 18, 16, 20, 20, 34, 40], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    info = wb.create_sheet("Yo‘riqnoma")
    for line in TEMPLATE_HELP:
        info.append([line])
    info.column_dimensions["A"].width = 82

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def template_csv() -> bytes:
    """Namunaviy .csv shablon (Excel uchun BOM bilan — kirill buzilmasin)."""
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", lineterminator="\r\n")
    w.writerow(TEMPLATE_HEADERS)
    for r in _template_rows():
        w.writerow(r)
    # Yo'riqnomani CSV ga qo'sha olmaymiz (bitta varaq) — izoh qatorlari
    # bo'lib yozamiz. Import sarlavhani qidirganda ularni tashlab ketadi.
    w.writerow([])
    for line in TEMPLATE_HELP:
        w.writerow(["# " + line if line else ""])
    return buf.getvalue().encode("utf-8-sig")


# --- To'ldirilgan shablonni o'qish ------------------------------------------
def _find_header(rows: List[List[Any]], limit: int = 10
                 ) -> Tuple[int, Dict[str, int], List[str]]:
    """Sarlavha qatorini topadi. Shart: `doc_type` va `name` tanilgan bo'lsin."""
    best: Optional[Tuple[int, Dict[str, int], List[str]]] = None
    for i, row in enumerate(rows[:limit]):
        if not row:
            continue
        mapping: Dict[str, int] = {}
        unknown: List[str] = []
        for idx, h in enumerate(row):
            if h is None or str(h).strip() == "":
                continue
            f = _match_column(h)
            if f and f not in mapping:
                mapping[f] = idx
            elif f is None:
                unknown.append(str(h).strip())
        if "doc_type" in mapping and "name" in mapping:
            if best is None or len(mapping) > len(best[1]):
                best = (i, mapping, unknown)
    if best is None:
        raise _importer().ImportFormatError(
            "Sarlavha qatori topilmadi: “Hujjat turi” va “Hujjat nomi” "
            "ustunlari bo‘lishi shart. Shablonni yuklab olib, o‘shani "
            "to‘ldiring.")
    return best


#: Shablondagi TO'LDIRILMAGAN qatorni aniqlash uchun: bu maydonlardan
#: hech biri bo'lmasa, qator "hali to'ldirilmagan" deb o'tkazib yuboriladi.
#: NEGA: shablon 11 ta qator bilan keladi. Ularni "bor" deb import qilsak,
#: cheklist hammasini YASHIL qilib qo'yardi — hujjat aslida yo'q bo'lsa ham.
#: Bu jimgina yolg'on bo'lardi, shuning uchun bo'sh qator import QILINMAYDI.
_FILLED_FIELDS = ("number", "issued_at", "valid_until", "file_name", "file_ref")


def _is_example_row(row: Sequence[Any], mapping: Dict[str, int]) -> bool:
    """Shablonning O'ZGARTIRILMAGAN misol qatorimi?

    Misol qator TO'LDIRILGAN holda keladi (raqam, sanalar, havola) — aks
    holda uni qanday to'ldirish kerakligi ko'rinmasdi. Lekin shu sababli u
    "to'ldirilgan qator" tekshiruvidan o'tib ketadi va foydalanuvchi uni
    o'chirishni unutsa, bazaga SOXTA hujjat yozilardi ("AA 1234567",
    "https://disk.example/…"). O'lchangan holat: shablon o'zgartirilmasdan
    yuklanganda bazada haqiqiy guvohnoma paydo bo'ldi.

    Shuning uchun BARCHA kataklari misol bilan bir xil bo'lgan qator
    o'tkazib yuboriladi. BITTA katakni tahrirlash yetarli — u holda qator
    foydalanuvchining ma'lumoti hisoblanadi va normal import bo'ladi.
    """
    for i, field in enumerate(f for f, _, _ in TEMPLATE_COLUMNS):
        idx = mapping.get(field)
        got = "" if idx is None or idx >= len(row) or row[idx] is None else str(row[idx])
        want = _EXAMPLE_ROW[i]
        # Sana Excel'da date obyektiga aylanib qolishi mumkin — ikkalasini
        # ham kanonik shaklga keltirib solishtiramiz.
        if field in ("issued_at", "valid_until"):
            d_got, _e, perp_got = parse_date(row[idx] if idx is not None
                                             and idx < len(row) else None)
            d_want, _e2, perp_want = parse_date(want)
            if (d_got, perp_got) != (d_want, perp_want):
                return False
        elif canon(got).strip() != canon(str(want)).strip():
            return False
    return True


def parse_document_rows(rows: List[List[Any]], mapping: Dict[str, int],
                        header_idx: int
                        ) -> Tuple[List[Dict[str, Any]], List[Dict], List[Dict]]:
    """Xom jadvaldan tozalangan hujjatlar + xatolar + ogohlantirishlar.

    `row` — FAYLDAGI qator raqami (Excel'da ko'rinadigani bilan bir xil).
    """
    imp = _importer()
    ok: List[Dict[str, Any]] = []
    errors: List[Dict] = []
    warnings: List[Dict] = []
    seen: Dict[Tuple[str, str], int] = {}

    def cell(row: List[Any], field: str) -> Any:
        idx = mapping.get(field)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return v if not isinstance(v, str) or v.strip() else None

    def err(row_no: int, field: Optional[str], value: Any, message: str) -> Dict:
        label = next((h for f, h, _ in TEMPLATE_COLUMNS if f == field), field or "—")
        return {"row": row_no, "column": label, "field": field,
                "value": None if value is None else str(value)[:120],
                "message": message}

    skipped: List[str] = []          # to'ldirilmagan shablon qatorlari
    example_skipped = False          # o'zgartirilmagan misol qator uchramadimi

    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        row_no = i + 1
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        # CSV shablonining oxiridagi yo'riqnoma qatorlari ("# ...") — izoh.
        # Ular xato ham, ogohlantirish ham emas: fayl shunday yasалgan.
        first = next((str(c).strip() for c in row if c is not None
                      and str(c).strip()), "")
        if first.startswith("#"):
            continue

        # O'zgartirilmagan misol qator — foydalanuvchining ma'lumoti emas
        if _is_example_row(row, mapping):
            example_skipped = True
            continue

        raw_type = cell(row, "doc_type")
        name = imp.cell_text(cell(row, "name"))

        # To'ldirilmagan shablon qatori — o'tkazamiz. Shablon 11 ta qator
        # bilan keladi, ularning ko'pi bo'sh qolishi ODATIY hol, shuning
        # uchun har biriga alohida ogohlantirish emas — oxirida BITTA
        # yig'ma xabar (aks holda hisobot shovqinga to'lardi).
        if not any(cell(row, f) is not None for f in _FILLED_FIELDS):
            skipped.append(imp.cell_text(raw_type) or name or f"{row_no}-qator")
            continue

        row_errors: List[Dict] = []

        doc_type = match_doc_type(raw_type)
        if not doc_type:
            row_errors.append(err(
                row_no, "doc_type", raw_type,
                "Hujjat turi tanilmadi. Shablondagi tayyor nomlardan birini "
                "qoldiring (masalan “Litsenziya / faoliyat ruxsatnomasi”)."))
        if not name:
            row_errors.append(err(row_no, "name", None,
                                  "Hujjat nomi bo‘sh — qator qabul qilinmadi."))
        elif len(name) > 300:
            row_errors.append(err(row_no, "name", name,
                                  "Nom juda uzun (300 belgidan ko‘p)."))

        issued, e_iss, _ = parse_date(cell(row, "issued_at"))
        if e_iss:
            row_errors.append(err(row_no, "issued_at", cell(row, "issued_at"),
                                  f"Berilgan sana — {e_iss}."))
        valid, e_val, perpetual = parse_date(cell(row, "valid_until"))
        if e_val:
            row_errors.append(err(row_no, "valid_until", cell(row, "valid_until"),
                                  f"Amal qilish muddati — {e_val}."))

        if issued and valid and valid < issued:
            row_errors.append(err(
                row_no, "valid_until", cell(row, "valid_until"),
                "Amal qilish muddati berilgan sanadan oldin — sanalarni "
                "tekshiring."))

        key = (doc_type or "", (name or "").lower())
        if not row_errors and key in seen:
            row_errors.append(err(
                row_no, "name", name,
                f"Bu tur va nom faylda takrorlangan ({seen[key]}-qatorda ham "
                f"bor). Nomini farqlang yoki qatorni o‘chiring."))

        if row_errors:
            errors.extend(row_errors)
            continue

        seen[key] = row_no

        # Muddat KO'RSATILMAGAN (bo'sh, "muddatsiz" ham yozilmagan) — cheklist
        # buni "cheklanmagan" deb o'qiydi va hujjat doim yaroqli ko'rinadi.
        # Jim qolmaymiz: bu foydalanuvchi kutmagan natija bo'lishi mumkin.
        if valid is None and not perpetual:
            warnings.append(err(
                row_no, "valid_until", None,
                "Amal qilish muddati ko‘rsatilmagan — hujjat MUDDATSIZ deb "
                "hisoblanadi va cheklistda hech qachon “tugagan” bo‘lmaydi."))

        ok.append({
            "row": row_no,
            "doc_type": doc_type,
            "label": BY_CODE[doc_type]["label"],
            "name": name,
            "number": imp.cell_text(cell(row, "number")),
            "issued_at": issued,
            "valid_until": valid,
            "file_name": imp.cell_text(cell(row, "file_name")),
            "file_ref": imp.cell_text(cell(row, "file_ref")),
            "note": imp.cell_text(cell(row, "note")),
        })

    if example_skipped:
        warnings.append({
            "row": 0, "column": "—", "field": None, "value": None,
            "message": ("Shablondagi namunaviy misol qator o‘tkazib yuborildi "
                        "(u o‘zgartirilmagan). Uni o‘chirib qo‘yishingiz "
                        "mumkin."),
        })

    if skipped:
        shown = ", ".join(skipped[:6]) + ("…" if len(skipped) > 6 else "")
        warnings.append({
            "row": 0, "column": "—", "field": None, "value": None,
            "message": (f"{len(skipped)} ta qator to‘ldirilmagan va o‘tkazib "
                        f"yuborildi ({shown}). Ular cheklistda “yo‘q” bo‘lib "
                        f"qoladi."),
        })

    return ok, errors, warnings


#: Qayta yuklashda nusxa ko'paymasin: tur + nom (harf registri farqsiz).
DOC_FIND_SQL = """
SELECT id FROM company_document
WHERE doc_type = %(doc_type)s AND lower(name) = lower(%(name)s)
  AND company_id = %(company_id)s
ORDER BY id LIMIT 1
"""

DOC_IMPORT_UPDATE_SQL = """
UPDATE company_document SET
    number       = COALESCE(%(number)s, number),
    issued_at    = COALESCE(%(issued_at)s, issued_at),
    -- Muddat ATAYIN COALESCE emas: shablonda bo'sh qoldirilgan "Amal qiladi"
    -- MUDDATSIZ degani. COALESCE bo'lsa eski (tugagan) sana qolib ketardi va
    -- foydalanuvchi yangilaganini ko'rmasdi.
    valid_until  = %(valid_until)s,
    file_name    = COALESCE(%(file_name)s, file_name),
    file_ref     = COALESCE(%(file_ref)s, file_ref),
    note         = COALESCE(%(note)s, note),
    updated_at   = now()
WHERE id = %(id)s AND company_id = %(company_id)s
"""


def parse_document_file(data: bytes, filename: str) -> Tuple[List[Dict[str, Any]],
                                                             Dict[str, Any]]:
    """To'ldirilgan shablonni O'QIYDI va tekshiradi — BAZAGA TEGMAYDI.

    Qaytaradi: (tozalangan qatorlar, hisobot). Hisobotda ustunlar tanilgani,
    xatolar, ogohlantirishlar va ko'rish uchun namuna bor.

    NEGA ALOHIDA: shablon va uning qoidalari (sarlavhalarni tanish, sana
    formatlari, hujjat turini aniqlash) SHU MODULDA yashaydi, lekin natijani
    KIM saqlashi har xil bo'lishi mumkin — kompaniya hujjatlari
    (`company_document`) yoki tashqi tizim (ERP mijoz korxonalari).
    Parser ikkinchi marta yozilmasligi uchun yozish qismidan ajratilgan.
    """
    imp = _importer()
    rows, fmt = imp.read_table(data, filename)
    if not rows:
        raise imp.ImportFormatError("Fayl bo‘sh.")

    header_idx, mapping, unknown = _find_header(rows)
    ok, errors, warnings = parse_document_rows(rows, mapping, header_idx)

    report: Dict[str, Any] = {
        "filename": filename,
        "format": fmt,
        "columns": {
            "detected": {
                next(h for f, h, _ in TEMPLATE_COLUMNS if f == field):
                    (rows[header_idx][i] if i < len(rows[header_idx]) else None)
                for field, i in mapping.items()},
            "unknown": unknown,
            "missing": [h for f, h, _ in TEMPLATE_COLUMNS
                        if f in ("doc_type", "name") and f not in mapping],
        },
        "header_row": header_idx + 1,
        "rows_total": len(ok) + len({e["row"] for e in errors}),
        "rows_ok": len(ok),
        "rows_error": len({e["row"] for e in errors}),
        "errors": errors,
        "warnings": warnings,
        "preview": [_import_preview(r) for r in ok[:50]],
    }
    return ok, report


def import_documents(data: bytes, filename: str, company_id: int, *,
                     dry_run: bool = True) -> Dict[str, Any]:
    """To'ldirilgan shablonni o'qiydi va KOMPANIYA hujjatlariga yozadi.

    dry_run=True  — faqat tekshiradi, bazaga HECH NARSA yozilmaydi.
    dry_run=False — to'g'ri qatorlar bitta tranzaksiyada yoziladi.

    Katalog importi (P0-4) bilan bir xil shartnoma: xato BITTA QATORNI
    to'xtatadi, importni emas.
    """
    from api import db

    ok, report = parse_document_file(data, filename)
    result: Dict[str, Any] = {"dry_run": dry_run, **report,
                              "inserted": 0, "updated": 0}

    if dry_run or not ok:
        if ok:
            result["inserted"], result["updated"] = _import_forecast(ok, company_id)
        return result

    inserted = updated = 0
    with db.get_conn() as conn:
        try:
            with conn.cursor() as cur:
                for r in ok:
                    params = {"company_id": company_id}
                    params.update({k: r[k] for k in
                                   ("doc_type", "name", "number", "issued_at",
                                    "valid_until", "file_name", "file_ref",
                                    "note")})
                    cur.execute(DOC_FIND_SQL, params)
                    found = cur.fetchone()
                    if found:
                        cur.execute(DOC_IMPORT_UPDATE_SQL,
                                    {**params, "id": found["id"]})
                        updated += 1
                    else:
                        cur.execute(DOC_INSERT_SQL, params)
                        inserted += 1
            conn.commit()
        except Exception:
            conn.rollback()
            raise

    result["inserted"] = inserted
    result["updated"] = updated
    return result


def _import_forecast(ok: List[Dict[str, Any]], company_id: int) -> Tuple[int, int]:
    """Dry-run uchun: nechtasi qo'shiladi / nechtasi yangilanadi."""
    from api import db
    try:
        rows = db.query("SELECT doc_type, lower(name) AS n FROM company_document "
                        "WHERE company_id = %(company_id)s",
                        {"company_id": company_id})
    except Exception:
        return len(ok), 0
    existing = {(r["doc_type"], r["n"]) for r in rows}
    upd = sum(1 for r in ok if (r["doc_type"], r["name"].lower()) in existing)
    return len(ok) - upd, upd


def rows_json(ok: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Tozalangan qatorlar -> JSON (sanalar ISO satr). Parser xizmat sifatida
    berilganda (`POST /company/documents/parse`) chaqiruvchi shu ro'yxatni
    oladi va O'Z bazasiga yozadi."""
    return [_import_preview(r) for r in ok]


def _import_preview(r: Dict[str, Any]) -> Dict[str, Any]:
    """Frontend jadvali uchun (date -> ISO satr)."""
    return {
        "row": r["row"],
        "doc_type": r["doc_type"],
        "label": r["label"],
        "name": r["name"],
        "number": r["number"],
        "issued_at": r["issued_at"].isoformat() if r["issued_at"] else None,
        "valid_until": r["valid_until"].isoformat() if r["valid_until"] else None,
        "status": doc_status({"valid_until": r["valid_until"]}),
        "file_ref": r["file_ref"],
    }
